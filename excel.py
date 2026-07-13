import os
from datetime import datetime

from openpyxl import Workbook, load_workbook


def save_to_excel(data):
    file_name = "requests.xlsx"

    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(["Дата", "Тип", "Имя", "Телефон", "Описание", "Объект"])
        wb.save(file_name)

    wb = load_workbook(file_name)
    ws = wb.active

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        data.get("type"),
        data.get("name"),
        data.get("phone"),
        data.get("text"),
        data.get("property"),
    ])

    wb.save(file_name)